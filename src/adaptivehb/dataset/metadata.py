"""Metadata loading for the dataset subsystem.

The metadata file is the central reference for the dataset (DATASET_SPEC Ch.9).
Both CSV/TSV and Excel (``.xlsx``/``.xlsm``/``.xls``) files are supported: CSV
parsing uses only the standard library so the dataset core stays dependency-light
and testable without pandas, while Excel loading uses ``openpyxl`` (the reporting
extra) and raises a clear error if it is unavailable.
"""

from __future__ import annotations

import csv
import importlib.util
from pathlib import Path

from adaptivehb.dataset.bmi import add_bmi
from adaptivehb.exceptions import DatasetError

# File suffixes handled by the Excel loader (everything else is treated as
# delimited text and read with the csv module).
_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}


class MetadataTable:
    """An in-memory view of the patient metadata CSV."""

    def __init__(self, rows: list[dict[str, str]], columns: list[str], id_column: str) -> None:
        """Initialize from parsed rows.

        Args:
            rows: Metadata rows as dictionaries.
            columns: Column names in file order.
            id_column: Name of the patient-identifier column.
        """
        self._rows = rows
        self._columns = columns
        self._id_column = id_column
        self._by_id: dict[str, dict[str, str]] = {}
        for row in rows:
            pid = (row.get(id_column) or "").strip()
            if pid and pid not in self._by_id:
                self._by_id[pid] = row

    @classmethod
    def load(cls, path: str | Path, id_column: str) -> MetadataTable:
        """Load a metadata file (CSV/TSV or Excel) from disk.

        Args:
            path: Path to the metadata file. ``.xlsx``/``.xlsm``/``.xls`` are read
                as Excel; any other extension is read as delimited text.
            id_column: Name of the patient-identifier column.

        Returns:
            A populated :class:`MetadataTable`.

        Raises:
            DatasetError: If the file is missing, empty, or (for Excel) openpyxl
                is not installed.
        """
        file_path = Path(path)
        if not file_path.is_file():
            raise DatasetError(f"Metadata file not found: {file_path}")
        if file_path.suffix.lower() in _EXCEL_SUFFIXES:
            columns, rows = _read_excel(file_path)
        else:
            columns, rows = _read_delimited(file_path)
        if not columns:
            raise DatasetError(f"Metadata file has no header: {file_path}")
        return cls(rows, columns, id_column)

    def add_column(self, name: str) -> None:
        """Register a column name if it is not already present (order-preserving)."""
        if name not in self._columns:
            self._columns.append(name)

    def derive_bmi(
        self,
        *,
        height_column: str,
        weight_column: str,
        bmi_column: str,
        height_unit: str = "cm",
    ) -> int:
        """Fill a blank/absent BMI column from height and weight, in place.

        Rows are mutated in place (the id index shares the same row objects), and
        the BMI column is registered so validation and statistics can see it.

        Returns:
            The number of rows whose BMI was newly computed.
        """
        filled = add_bmi(
            self._rows,
            height_column=height_column,
            weight_column=weight_column,
            bmi_column=bmi_column,
            height_unit=height_unit,
        )
        if filled:
            self.add_column(bmi_column)
        return filled

    @property
    def columns(self) -> list[str]:
        """Column names in file order."""
        return list(self._columns)

    @property
    def rows(self) -> list[dict[str, str]]:
        """All metadata rows."""
        return list(self._rows)

    @property
    def patient_ids(self) -> list[str]:
        """Unique patient identifiers, in first-seen order."""
        return list(self._by_id.keys())

    def has_columns(self, names: list[str] | tuple[str, ...]) -> list[str]:
        """Return the subset of ``names`` that are missing from the table."""
        return [name for name in names if name not in self._columns]

    def get(self, patient_id: str) -> dict[str, str] | None:
        """Return the metadata row for a patient, or ``None`` if absent."""
        return self._by_id.get(patient_id)

    def __len__(self) -> int:
        """Number of rows in the raw table (including any duplicates)."""
        return len(self._rows)


def _read_delimited(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """Read a CSV/TSV file into ``(columns, rows)`` using the standard library."""
    delimiter = "\t" if path.suffix.lower() in {".tsv", ".tab"} else ","
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        columns = list(reader.fieldnames or [])
        rows = [dict(row) for row in reader]
    return columns, rows


def _read_excel(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """Read the first worksheet of an Excel file into ``(columns, rows)``.

    Requires ``openpyxl``. Cells are stringified so the rest of the pipeline (which
    treats metadata as text) is unchanged; blank cells become empty strings.
    """
    if importlib.util.find_spec("openpyxl") is None:
        raise DatasetError(
            f"Reading Excel metadata ({path.name}) requires 'openpyxl'. Install the "
            "reporting extra (pip install openpyxl) or export the file to CSV."
        )
    from openpyxl import load_workbook  # local import: optional dependency

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return [], []
        columns = [str(cell).strip() if cell is not None else "" for cell in header]
        rows: list[dict[str, str]] = []
        for raw in rows_iter:
            if raw is None or all(cell is None for cell in raw):
                continue  # skip fully blank rows
            row = {
                columns[i]: ("" if value is None else str(value).strip())
                for i, value in enumerate(raw)
                if i < len(columns) and columns[i]
            }
            rows.append(row)
        return [c for c in columns if c], rows
    finally:
        workbook.close()


__all__ = ["MetadataTable"]
